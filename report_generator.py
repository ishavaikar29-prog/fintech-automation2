import os
import smtplib
import logging
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders


# ============= Logging Setup =============

LOG_FILE = "error.log"

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.ERROR,
    format="%(asctime)s [%(levelname)s] %(message)s",
)


def log_error(context: str, exc: Exception):
    """Helper to log errors with some context."""
    logging.error("%s: %s", context, str(exc))


# ============= API Helpers =============

def fetch_api_data(url: str, headers=None):
    """Call an API and return JSON list. On error, log and return empty list."""
    try:
        print(f"Calling API: {url}")
        res = requests.get(url, headers=headers, timeout=20)
        res.raise_for_status()
        data = res.json()
        # Make sure it's always a list for easier handling
        if isinstance(data, dict):
            return [data]
        return data
    except Exception as e:
        log_error(f"API call failed for {url}", e)
        print(f"ERROR: API call failed for {url}, check {LOG_FILE}")
        return []


# ============= Excel Helpers =============

def style_header(row):
    """Apply bold font + background color to header row."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for cell in row:
        cell.font = header_font
        cell.fill = header_fill


def autosize_columns(ws):
    """Automatically adjust column width based on max length."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def write_sheet_from_json(ws, data, field_mapping):
    """
    data: list of dicts from API
    field_mapping: Ordered dict-like (header -> key_in_json)
    """
    # Header
    headers = list(field_mapping.keys())
    ws.append(headers)
    style_header(ws[1])

    # Rows
    for item in data:
        row = []
        for key in field_mapping.values():
            row.append(item.get(key))
        ws.append(row)

    autosize_columns(ws)


def create_report_excel(users, posts, todos, file_path="daily_report.xlsx"):
    """Create an Excel with 3 sheets: Users, Posts, Todos."""
    wb = Workbook()

    # Sheet 1: Users
    ws_users = wb.active
    ws_users.title = "Users"
    user_mapping = {
        "ID": "id",
        "Name": "name",
        "Username": "username",
        "Email": "email",
    }
    write_sheet_from_json(ws_users, users, user_mapping)

    # Sheet 2: Posts
    ws_posts = wb.create_sheet("Posts")
    post_mapping = {
        "ID": "id",
        "User ID": "userId",
        "Title": "title",
    }
    write_sheet_from_json(ws_posts, posts, post_mapping)

    # Sheet 3: Todos
    ws_todos = wb.create_sheet("Todos")
    todo_mapping = {
        "ID": "id",
        "User ID": "userId",
        "Title": "title",
        "Completed": "completed",
    }
    write_sheet_from_json(ws_todos, todos, todo_mapping)

    wb.save(file_path)
    return file_path


# ============= Email Helper =============

def send_email_with_attachments(
    smtp_host,
    smtp_port,
    smtp_user,
    smtp_pass,
    to_email,
    subject,
    body,
    attachments
):
    """Send email with multiple attachments."""
    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = to_email
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain"))

    for file_path in attachments:
        if not os.path.exists(file_path):
            continue

        part = MIMEBase("application", "octet-stream")
        with open(file_path, "rb") as f:
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(file_path)}"
        )
        msg.attach(part)

    try:
        server = smtplib.SMTP(smtp_host, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
        server.quit()
        print("Email sent successfully!")
    except Exception as e:
        log_error("Email sending failed", e)
        print("ERROR: Email sending failed, check error.log")

# ============= Main Flow =============

if __name__ == "__main__":
    print("Starting multi-API report job...")

    # Read email config from environment / GitHub Secrets
    SMTP_HOST = os.getenv("SMTP_HOST")
    SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
    SMTP_USER = os.getenv("SMTP_USER")
    SMTP_PASS = os.getenv("SMTP_PASS")
    TO_EMAIL = os.getenv("TO_EMAIL")

    # API URLs (for real use – set via secrets; here we give defaults)
    API1_URL = os.getenv("API1_URL", "https://jsonplaceholder.typicode.com/users")
    API2_URL = os.getenv("API2_URL", "https://jsonplaceholder.typicode.com/posts")
    API3_URL = os.getenv("API3_URL", "https://jsonplaceholder.typicode.com/todos")

    # Optional: common headers if your real APIs need API key
    API_KEY = os.getenv("API_KEY")  # can be None for public APIs
    headers = {}
    if API_KEY:
        headers["Authorization"] = f"Bearer {API_KEY}"

    # Step 1: Fetch data from multiple APIs
    users = fetch_api_data(API1_URL, headers=headers)
    posts = fetch_api_data(API2_URL, headers=headers)
    todos = fetch_api_data(API3_URL, headers=headers)

    # Step 2: Generate Excel report
    print("Creating Excel report...")
    report_file = create_report_excel(users, posts, todos)

    # Step 3: Compose email body (mention counts + error info)
    body_lines = [
        "Hello,",
        "",
        "Here is your automated multi-API report.",
        "",
        f"Users records: {len(users)}",
        f"Posts records: {len(posts)}",
        f"Todos records: {len(todos)}",
    ]

    if os.path.exists(LOG_FILE) and os.path.getsize(LOG_FILE) > 0:
        body_lines.append("")
        body_lines.append("⚠ Some errors were logged during execution. Please check attached system logs on the server / CI artifacts.")

    body_lines.append("")
    body_lines.append(f"Generated at: {datetime.utcnow().isoformat()} UTC")
    body_lines.append("")
    body_lines.append("-- Automation Bot")

    email_body = "\n".join(body_lines)

    # Step 4: Send email
    print("Sending email...")
    attachments = [report_file]

# If errors exist, attach error.log
    if os.path.exists(LOG_FILE) and os.path.getsize(LOG_FILE) > 0:
        attachments.append(LOG_FILE)
        
    send_email_with_attachments(
        smtp_host=SMTP_HOST,
        smtp_port=SMTP_PORT,
        smtp_user=SMTP_USER,
        smtp_pass=SMTP_PASS,
        to_email=TO_EMAIL,
        subject="Daily Multi-API Report",
        body=email_body,
        attachments=attachments,
)

    print("Job finished.")
